from openpyxl import load_workbook
from openpyxl import Workbook

from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

import os


#Configurações gráficas
#----------------------

celulaAmarela = PatternFill(start_color='00FFFFCC',
                               end_color='00FFFFCC',
                               fill_type='solid')

bordaFinaPreta = Side(border_style="thin", color="000000")
bordasPlanilha = Border(left=bordaFinaPreta, 
                        right=bordaFinaPreta,
                        top=bordaFinaPreta,
                        bottom=bordaFinaPreta)


#Caminhos de arquivo & variáveis chave
#-------------------------------------

caminho_arquivo = "C:\\Users\\PC\\Downloads\\algorithm-quickstart\\jupyter\\scripts\\Planilha.xlsx"#Determina o caminho do arquivo
planilha = load_workbook(filename=caminho_arquivo)#Abre o arquivo
guia_selecionada = planilha['Guia']#Determina guia_selecionada como a guia Guia da planilha

criar_novo_arquivo = Workbook()
nome_novo = ""


total_linhas = len(guia_selecionada['A']) + 1#Retorna o total de linhas com conteúdo

for linha in range(2, len(guia_selecionada['A']) + 1):#Cria um loop a partir da segunda linha até a última com conteúdo
    
    nome_atual = guia_selecionada['A%s' % linha].value#Armazena o valor indexado da linha atual a cada loop
    
    if nome_novo == nome_atual:
        
        linha_guia_quebra = len(selecionar_nova_planilha['A']) + 1#Obtém a quantidade de linhas
        celula_colunaA = "A" + str(linha_guia_quebra)#Concatena o valor da linha (1, 2, etc) com A formando A1, A2, etc
        celula_colunaB = "B" + str(linha_guia_quebra)#Concatena o valor da linha (1, 2, etc) com B formando B1, B2, etc
        celula_colunaC = "C" + str(linha_guia_quebra)#Concatena o valor da linha (1, 2, etc) com C formando C1, C2, etc
        
        selecionar_nova_planilha[celula_colunaA] = guia_selecionada['A%s' % linha].value#Indexa os dados da coluna A da guia original
        selecionar_nova_planilha[celula_colunaB] = guia_selecionada['B%s' % linha].value#Indexa os dados da coluna B da guia original
        selecionar_nova_planilha[celula_colunaC] = guia_selecionada['C%s' % linha].value#Indexa os dados da coluna C da guia original
        
        #--------------------------------
        #Aplicando configurações gráficas
        
        selecionar_nova_planilha[celula_colunaA].fill = celulaAmarela
        selecionar_nova_planilha[celula_colunaB].fill = celulaAmarela
        selecionar_nova_planilha[celula_colunaC].fill = celulaAmarela
        #--------------------------------
        
        criar_nova_planilha.save(filename=caminho_nova_planilha)
        
        
    else:
        
        nome_novo = guia_selecionada['A%s' % linha].value
        
        nova_planilha = criar_novo_arquivo.active
        nova_planilha.title = "NovaGuia"
        
        #Determina o caminho da nova planilha
        caminho_nova_planilha = "C:\\Users\\PC\\Downloads\\algorithm-quickstart\\jupyter\\scripts\\" + guia_selecionada['A%s' % linha].value + ".xlsx"
        selecionar_nova_planilha = criar_novo_arquivo['NovaGuia']
        
        
        selecionar_nova_planilha['A1'] = "Título 1"#Cria um titulo para a nova guia
        selecionar_nova_planilha['B1'] = "Título 2"#Cria outro titulo para a nova guia
        selecionar_nova_planilha['C1'] = "Título 3"#Cria um último título para a nova guia
        
        selecionar_nova_planilha['A2'] = guia_selecionada['A%s' % linha].value#Indexa os dados unitários da coluna A da outra guia na nova
        selecionar_nova_planilha['B2'] = guia_selecionada['B%s' % linha].value#Indexa os dados unitários da coluna B da outra guia na nova
        selecionar_nova_planilha['C2'] = guia_selecionada['C%s' % linha].value#Indexa os dados unitários da coluna C da outra guia na nova
        
        #----------------------------------
        #Configurando o tamanho das colunas
        
        selecionar_nova_planilha.column_dimensions["A"].width = 18
        selecionar_nova_planilha.column_dimensions["B"].width = 25
        selecionar_nova_planilha.column_dimensions["C"].width = 15
        #----------------------------------
        
        selecionar_nova_planilha.delete_rows(3, 100)#Deleta as linhas a partir da 3ª ate a 100ª
        criar_nova_planilha.save(filename=caminho_nova_planilha)